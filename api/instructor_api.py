"""Instructor-facing API endpoints that back the list, detail, and setup screens.

These sit alongside the existing InstructorSimulationsView / InstructorDeployStudentsView
in views.py and follow the same conventions: JWT auth, the IsInstructor permission, and
every cohort scoped to the instructor who teaches it. The frontend speaks in "rounds" and
"firms"; underneath those are DigitalCo weeks and teams, and the mapping happens here.

Nothing in this module touches the scoring engine. Advancing a round only moves each run's
week pointer forward; scoring still flows through the engine as students submit.
"""
import calendar
import csv
import io
import secrets
from datetime import date, timedelta

from django.conf import settings
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.db.models import Count, Q, Sum
from django.db.models.functions import Coalesce
from django.utils import timezone as dj_timezone
from django.utils.dateparse import parse_date
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from advisors.models import BILLED, AdvisorSession
from mailer.accounts import send_faculty_invite
from mailer.invites import invite_url, send_invitation
from core.models import (
    Cohort, Enrollment, Invitation, InvitationStatus, Run, RunStatus, Team, User, UserRole,
)
from weeks.models import WeekInstanceStatus

from .permissions import IsInstructor
from .views import ADMIN_TOTAL_ROUNDS, _instructor_cohort_row

TEST_PASSWORD = 'test1234'


# ---- shared helpers -------------------------------------------------------

def _cohort_for(request, cohort_id):
    """A cohort this instructor teaches, or 404. Keeps one instructor out of another's sim."""
    return get_object_or_404(Cohort, id=cohort_id, instructors=request.user)


def _firm_numbering(cohort):
    """Stable firm numbers for a cohort's teams. Returns (ordered_teams, {team_id: number})."""
    teams = list(Team.objects.filter(cohort=cohort).order_by('id'))
    numbering = {t.id: i + 1 for i, t in enumerate(teams)}
    return teams, numbering


def _registration_url(request, token):
    if not token:
        return None
    base = getattr(settings, 'FRONTEND_BASE_URL', '') or request.build_absolute_uri('/')[:-1]
    return f"{base.rstrip('/')}/register/{token}"


def _slug(text):
    """A filename-safe stem, so a cohort called 'MIS 7000 / Fall' cannot produce
    a path separator or a quote inside a Content-Disposition header."""
    import re

    return re.sub(r'[^A-Za-z0-9]+', '-', str(text or '')).strip('-').lower()[:60]


def _csv_safe(value):
    """Defuse a cell Excel would treat as a formula.

    Invitation emails come from whatever an instructor typed or uploaded, so a
    value beginning =, +, - or @ is reachable by a hostile roster. Excel and
    Sheets execute those on open; prefixing an apostrophe keeps the text intact
    and inert. Links are unaffected because they start with 'h'.
    """
    text = '' if value is None else str(value)
    return "'" + text if text[:1] in ('=', '+', '-', '@') else text


def _fmt_date(d):
    return f'{calendar.month_abbr[d.month]} {d.day}, {d.year}'


def _rounds(cohort, current_round):
    """Per-round schedule, honoring any stored extensions. Extending a round lengthens it and
    pushes every later round forward, which mirrors what the detail screen promises.

    The active round is re-anchored to when it actually opened, and everything
    after it follows from there. Without that, the whole calendar is
    `start_date + N x length`, which is only correct if rounds advance on exactly
    that cadence — they do not. An instructor who advances a day early leaves R2
    apparently closing in fourteen days while R1 showed seven, and the error
    compounds every round. Completed rounds keep their original dates: they are
    history, and rewriting them would be a different lie.
    """
    total = ADMIN_TOTAL_ROUNDS
    base_days = cohort.days_per_week or 7
    ext = cohort.round_extensions or {}
    opened = getattr(cohort, 'current_round_opened_at', None)
    opened_date = opened.date() if opened else None

    rows = []
    cursor = cohort.start_date
    for n in range(1, total + 1):
        extra = int(ext.get(str(n), 0))
        length = base_days + extra
        # Re-anchor at the active round, then let later rounds run on from it.
        if n == current_round and opened_date:
            cursor = opened_date
        if cursor:
            s = cursor
            e = cursor + timedelta(days=length)
            start_str, end_str = _fmt_date(s), _fmt_date(e)
            cursor = e
        else:
            start_str = end_str = '\u2014'
        status = 'Completed' if n < current_round else 'Active' if n == current_round else 'Upcoming'
        rows.append({'n': n, 'start': start_str, 'end': end_str, 'status': status, 'extended_days': extra})
    return rows


# A war-room hour costs several times a 1:1 hour, so faculty need the two split
# out, not just a combined total they can't explain to a student who queries it.
GROUP_HOURS = Count('id', filter=Q(group_session__isnull=False))
GROUP_BILLED = Sum(BILLED, filter=Q(group_session__isnull=False))


def _advisor_usage_by_student(cohort):
    """{student_id: {'hours', 'due', 'group_hours', 'group_due'}} for the cohort.

    hours/due are the totals across both modes; group_* is the war-room slice of
    them, so 1:1 usage is the difference.
    """
    rows = (
        AdvisorSession.objects
        .for_cohort(cohort)
        .values('student_id')
        .annotate(
            hours=Count('id'), due=Sum(BILLED),
            group_hours=GROUP_HOURS, group_due=GROUP_BILLED,
        )
    )
    return {
        r['student_id']: {
            'hours': r['hours'],
            'due': r['due'] or 0,
            'group_hours': r['group_hours'] or 0,
            'group_due': r['group_due'] or 0,
        }
        for r in rows
    }


def _student_rows(enrollments, numbering, advisor_usage=None):
    advisor_usage = advisor_usage or {}
    rows = []
    for e in enrollments:
        team = e.team
        number = numbering.get(team.id) if team else None
        usage = advisor_usage.get(e.student_id, {})
        rows.append({
            'enrollment_id': e.id,
            'id': e.student_id,
            'name': (e.student.get_full_name() or e.student.username),
            'email': e.student.email,
            'firm': team.name if team else None,
            'firm_index': (number - 1) if number else None,
            'paid': e.paid,
            'advisor_hours': usage.get('hours', 0),
            'advisor_due': usage.get('due', 0),
            'group_hours': usage.get('group_hours', 0),
            'group_due': usage.get('group_due', 0),
        })
    rows.sort(key=lambda r: ((r['firm_index'] if r['firm_index'] is not None else 999), r['name'].lower()))
    return rows


def _billing(cohort, enrollments):
    price = cohort.price_per_student or 0
    total = len(enrollments)
    paid = sum(1 for e in enrollments if e.paid)
    advisor = AdvisorSession.objects.for_cohort(cohort).aggregate(
        hours=Count('id'), billed=Sum(BILLED),
        group_hours=GROUP_HOURS, group_billed=GROUP_BILLED,
    )
    advisor_hours = advisor['hours'] or 0
    advisor_billed = advisor['billed'] or 0
    return {
        'price_per_student': price,
        'advisor_hourly_rate': cohort.advisor_hourly_rate or 0,
        'advisor_hours': advisor_hours,
        'advisor_billed': advisor_billed,
        # The war-room slice of the two figures above.
        'group_hours': advisor['group_hours'] or 0,
        'group_billed': advisor['group_billed'] or 0,
        'total_billed': total * price + advisor_billed,
        'received': paid * price,
        'pending': (total - paid) * price + advisor_billed,
        'paid_count': paid,
        'total_count': total,
    }


def _sync_team_membership(cohort, student, team):
    """Keep Team.members in step with an enrollment's firm so the existing engine and web
    flows (which read team.members) stay correct. A student sits in exactly one firm."""
    for t in Team.objects.filter(cohort=cohort).exclude(id=getattr(team, 'id', None)):
        t.members.remove(student)
    if team is not None:
        team.members.add(student)


# ---- detail ---------------------------------------------------------------

class InstructorSimulationDetailView(APIView):
    """Everything the detail and setup screens read for one simulation, plus the
    handful of settings faculty can change mid-run."""

    permission_classes = [IsAuthenticated, IsInstructor]

    def patch(self, request, cohort_id):
        """Settings faculty can change mid-run: the advisor rate and the start date.

        Both were set once at provisioning and stuck there, which left faculty
        unable to correct a wrong figure or a term that slipped a week.

        advisor_hourly_rate — hours already billed keep the rate snapshotted on
        their AdvisorSession, so this prices the *next* hour and never rewrites
        history.

        start_date — the round calendar is derived from it on every read (see
        _rounds), never stored, so moving the date regenerates the whole schedule
        by itself. Extensions already granted are keyed by round number and
        survive the move, shifting along with the rounds they belong to.
        """
        cohort = _cohort_for(request, cohort_id)

        fields = ('advisor_hourly_rate', 'start_date')
        if not any(f in request.data for f in fields):
            return Response({'detail': 'Nothing to update.'}, status=400)

        errors = {}
        changed = []
        previous_rate = cohort.advisor_hourly_rate or 0
        previous_start = cohort.start_date

        if 'advisor_hourly_rate' in request.data:
            try:
                rate = int(request.data.get('advisor_hourly_rate'))
            except (TypeError, ValueError):
                errors['advisor_hourly_rate'] = 'The advisor rate must be a whole number.'
            else:
                if not 0 <= rate <= 100000:
                    errors['advisor_hourly_rate'] = 'The advisor rate must be between 0 and 100000.'
                else:
                    cohort.advisor_hourly_rate = rate
                    changed.append('advisor_hourly_rate')

        if 'start_date' in request.data:
            raw = (request.data.get('start_date') or '').strip()
            if not raw:
                errors['start_date'] = 'Set a start date — the round calendar is built from it.'
            else:
                # parse_date returns None for a malformed string but *raises*
                # for a well-formed impossible one ("2026-13-45").
                try:
                    parsed = parse_date(raw)
                except ValueError:
                    parsed = None
                if parsed is None:
                    errors['start_date'] = 'That start date is not a valid date.'
                else:
                    cohort.start_date = parsed
                    changed.append('start_date')
                    # Moving the start date is an explicit re-plan of the whole
                    # calendar. A stale open-time anchor would quietly override
                    # the new dates for the round in play, so it goes.
                    cohort.current_round_opened_at = None
                    changed.append('current_round_opened_at')

        if errors:
            return Response(
                {'detail': 'Some fields need attention.', 'errors': errors}, status=400,
            )

        cohort.save(update_fields=changed)

        runs = Run.objects.filter(team__cohort=cohort)
        current_round = max((r.current_week for r in runs), default=1) or 1
        enrollments = list(Enrollment.objects.filter(cohort=cohort))
        return Response({
            'advisor_hourly_rate': cohort.advisor_hourly_rate,
            'previous_advisor_hourly_rate': previous_rate,
            'start_date': cohort.start_date.isoformat() if cohort.start_date else None,
            'previous_start_date': previous_start.isoformat() if previous_start else None,
            # The regenerated calendar, so the caller can render it without a refetch.
            'rounds': _rounds(cohort, current_round),
            'billing': _billing(cohort, enrollments),
        })

    def get(self, request, cohort_id):
        cohort = _cohort_for(request, cohort_id)
        teams, numbering = _firm_numbering(cohort)
        enrollments = list(
            Enrollment.objects.filter(cohort=cohort).select_related('student', 'team')
        )
        runs = Run.objects.filter(team__cohort=cohort)
        current_round = max((r.current_week for r in runs), default=1) or 1
        # True once every team's run is finalized. Lets the UI offer "Complete
        # simulation" at round 14 (which advance-round performs) and then show a
        # finished state, instead of leaving the advance control dead at the end.
        completed = runs.exists() and not runs.exclude(status=RunStatus.COMPLETE).exists()

        return Response({
            'id': cohort.id,
            'name': cohort.name,
            'tier': cohort.tier,
            'timezone': cohort.timezone,
            'start_date': cohort.start_date.isoformat() if cohort.start_date else None,
            'days_per_round': cohort.days_per_week,
            'enrollment_capacity': cohort.enrollment_capacity,
            'price_per_student': cohort.price_per_student,
            'advisor_hourly_rate': cohort.advisor_hourly_rate,
            'deployment_status': cohort.deployment_status,
            'current_round': current_round,
            'total_rounds': ADMIN_TOTAL_ROUNDS,
            'completed': completed,
            'registration_url': _registration_url(request, cohort.registration_token),
            # Whether this server can actually send email. The console backend
            # "succeeds" — it prints — so an invitation records as delivered
            # while nothing leaves the building. Faculty need to be told that
            # once, at the server level, rather than inferring it per row.
            'mail': {
                'backend': getattr(settings, 'MAIL_BACKEND', 'console'),
                'configured': bool(
                    getattr(settings, 'MAIL_BACKEND', '') == 'mailjet'
                    and getattr(settings, 'MAILJET_API_KEY', '')
                    and getattr(settings, 'MAILJET_API_SECRET', '')
                    and getattr(settings, 'MAIL_FROM_ADDRESS', '')
                ),
            },
            'instructors': [
                {
                    'id': i.id,
                    'first_name': i.first_name,
                    'last_name': i.last_name,
                    'username': i.username,
                    'email': i.email,
                }
                for i in cohort.instructors.all()
            ],
            'firms': [{'id': t.id, 'number': numbering[t.id], 'name': t.name} for t in teams],
            'students': _student_rows(enrollments, numbering, _advisor_usage_by_student(cohort)),
            'billing': _billing(cohort, enrollments),
            'rounds': _rounds(cohort, current_round),
        })


# ---- advance round --------------------------------------------------------

class InstructorAdvanceRoundView(APIView):
    """Move every team in the cohort to the next round. Instructor override: it doesn't wait
    for submissions and doesn't score (scoring stays in the engine)."""

    permission_classes = [IsAuthenticated, IsInstructor]

    def post(self, request, cohort_id):
        cohort = _cohort_for(request, cohort_id)
        runs = list(Run.objects.filter(team__cohort=cohort))
        advanced = 0
        for run in runs:
            if run.status == RunStatus.COMPLETE:
                continue
            if run.current_week >= ADMIN_TOTAL_ROUNDS:
                run.status = RunStatus.COMPLETE
                run.save(update_fields=['status'])
            else:
                run.current_week += 1
                run.save(update_fields=['current_week'])
                advanced += 1
        current = max((r.current_week for r in Run.objects.filter(team__cohort=cohort)), default=1)
        if advanced:
            # Stamp when this round actually opened so the countdown measures
            # from now rather than from a calendar slot the cohort may be
            # nowhere near.
            cohort.current_round_opened_at = dj_timezone.now()
            cohort.save(update_fields=['current_round_opened_at'])
        return Response({'advanced': advanced > 0, 'current_round': current})


# ---- invitations ----------------------------------------------------------

def _invite_json(inv):
    return {
        'id': str(inv.id),
        'email': inv.email,
        'status': inv.status,
        # An instructor needs to see who actually received the mail, not just
        # who was added to the list.
        'sent_at': inv.sent_at.isoformat() if inv.sent_at else None,
        'send_error': inv.send_error,
        # The exact link that was emailed, so an instructor can check it works,
        # or hand it over another way when mail fails or lands in spam. Only for
        # invitations still outstanding: a redeemed token is spent, and offering
        # it to copy invites a support conversation about a link that cannot
        # work. Instructor-only, and only for cohorts they teach.
        'url': invite_url(inv.token) if inv.status == InvitationStatus.PENDING else None,
    }


class InstructorInvitationsView(APIView):
    permission_classes = [IsAuthenticated, IsInstructor]

    def get(self, request, cohort_id):
        cohort = _cohort_for(request, cohort_id)
        invites = Invitation.objects.filter(cohort=cohort)
        return Response([_invite_json(i) for i in invites])


class InstructorInviteExportView(APIView):
    """Download the outstanding invitations as a spreadsheet.

    Email and link, one row per student, so an instructor can mail-merge them,
    hand them out in class, or paste a link for the one person whose mail
    bounced. Only invitations still pending: a redeemed token is spent, and a
    sheet full of dead links is worse than no sheet.

    Emits a real .xlsx when openpyxl is installed and falls back to CSV when it
    is not, matching what the bulk-invite importer already does. Both open in
    Excel, so no deployment has to grow a dependency for this.
    """

    permission_classes = [IsAuthenticated, IsInstructor]

    HEADERS = ('Email', 'Invitation link', 'Firm', 'Invited', 'Emailed')

    def get(self, request, cohort_id):
        cohort = _cohort_for(request, cohort_id)
        invites = (
            Invitation.objects
            .filter(cohort=cohort, status=InvitationStatus.PENDING)
            .select_related('team')
            .order_by('email')
        )
        rows = [
            (
                inv.email,
                invite_url(inv.token),
                inv.team.name if inv.team else '',
                _fmt_date(inv.created_at.date()) if inv.created_at else '',
                _fmt_date(inv.sent_at.date()) if inv.sent_at else 'not sent',
            )
            for inv in invites
        ]

        stem = _slug(cohort.name) or 'cohort'
        try:
            content, content_type, ext = self._xlsx(cohort, rows)
        except ImportError:
            content, content_type, ext = self._csv(rows)

        response = HttpResponse(content, content_type=content_type)
        response['Content-Disposition'] = (
            f'attachment; filename="{stem}-pending-invitations.{ext}"'
        )
        return response

    def _xlsx(self, cohort, rows):
        from io import BytesIO

        from openpyxl import Workbook  # raises ImportError when unavailable
        from openpyxl.styles import Font

        book = Workbook()
        sheet = book.active
        sheet.title = 'Pending invitations'
        sheet.append(list(self.HEADERS))
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in rows:
            sheet.append(list(row))
        for column, width in zip('ABCDE', (34, 78, 16, 14, 14)):
            sheet.column_dimensions[column].width = width
        sheet.freeze_panes = 'A2'

        buffer = BytesIO()
        book.save(buffer)
        return (
            buffer.getvalue(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'xlsx',
        )

    def _csv(self, rows):
        import csv
        from io import StringIO

        buffer = StringIO()
        writer = csv.writer(buffer)
        writer.writerow(self.HEADERS)
        for row in rows:
            writer.writerow([_csv_safe(value) for value in row])
        # BOM so Excel opens UTF-8 addresses correctly on a double click rather
        # than mangling any non-ASCII name in them.
        return '\ufeff' + buffer.getvalue(), 'text/csv; charset=utf-8', 'csv'


class InstructorInviteView(APIView):
    """Invite one student by email. Idempotent per (cohort, email)."""

    permission_classes = [IsAuthenticated, IsInstructor]

    def post(self, request, cohort_id):
        cohort = _cohort_for(request, cohort_id)
        email = (request.data.get('email') or '').strip().lower()
        if not email or '@' not in email:
            return Response({'detail': 'A valid email is required.'}, status=400)
        inv, created = Invitation.objects.get_or_create(
            cohort=cohort,
            email=email,
            defaults={'token': secrets.token_urlsafe(24), 'invited_by': request.user},
        )
        # Send on create, and on re-invite of an unsent one — an instructor
        # clicking invite twice means "this person still hasn't got it".
        sent, detail = (False, 'Already sent.')
        if created or not inv.sent_at:
            sent, detail = send_invitation(inv)
        body = _invite_json(inv)
        body['sent'] = bool(inv.sent_at)
        if not sent and not inv.sent_at:
            body['send_error'] = detail
        return Response(body, status=201 if created else 200)


class InstructorBulkInviteView(APIView):
    """Invite many students from an uploaded .csv (or .xlsx if openpyxl is installed)."""

    permission_classes = [IsAuthenticated, IsInstructor]

    def post(self, request, cohort_id):
        cohort = _cohort_for(request, cohort_id)
        upload = request.FILES.get('file')
        if not upload:
            return Response({'detail': 'Attach a file under the "file" field.'}, status=400)

        try:
            rows = _read_emails(upload)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=400)

        invited, skipped, errors = [], [], []
        seen = set()
        for row_no, raw in rows:
            email = (raw or '').strip().lower()
            if not email or '@' not in email:
                errors.append({'row': row_no, 'email': raw or '', 'detail': 'Not a valid email'})
                continue
            if email in seen:
                skipped.append({'row': row_no, 'email': email, 'detail': 'Duplicate in file'})
                continue
            seen.add(email)
            invitation, created = Invitation.objects.get_or_create(
                cohort=cohort,
                email=email,
                defaults={'token': secrets.token_urlsafe(24), 'invited_by': request.user},
            )
            if created:
                sent, detail = send_invitation(invitation)
                invited.append({
                    'row': row_no, 'email': email, 'created': True,
                    'sent': sent, **({'send_error': detail} if not sent else {}),
                })
            else:
                skipped.append({'row': row_no, 'email': email, 'detail': 'Already invited'})

        return Response({
            'summary': {
                'invited': len(invited),
                'sent': sum(1 for r in invited if r.get('sent')),
                'skipped': len(skipped),
                'errors': len(errors),
                'rows': len(rows),
            },
            'invited': invited,
            'skipped': skipped,
            'errors': errors,
        })


class InstructorInviteResendView(APIView):
    """Send an outstanding invitation again.

    The token is deliberately NOT rotated. Resending exists because a student
    says "I never got it", and rotating would kill every copy already sitting in
    their inbox — including the one they are about to find in spam. An instructor
    resending twice while chasing someone would leave only the newest mail
    working, and every earlier link reading "this link is not valid".

    Pass {"reissue": true} to deliberately retire the old links — for a mistyped
    address, or an invite forwarded to the wrong person.
    """

    permission_classes = [IsAuthenticated, IsInstructor]

    def post(self, request, cohort_id, invitation_id):
        cohort = _cohort_for(request, cohort_id)
        invitation = get_object_or_404(Invitation, id=invitation_id, cohort=cohort)
        if invitation.status != InvitationStatus.PENDING:
            return Response(
                {'detail': 'That invitation has already been accepted.'}, status=409,
            )
        if request.data.get('reissue'):
            invitation.token = secrets.token_urlsafe(24)
            invitation.save(update_fields=['token'])
        sent, detail = send_invitation(invitation)
        if not sent:
            return Response({'detail': detail, **_invite_json(invitation)}, status=502)
        return Response(_invite_json(invitation))


class InstructorBulkInviteTemplateView(APIView):
    permission_classes = [IsAuthenticated, IsInstructor]

    def get(self, request, cohort_id):
        _cohort_for(request, cohort_id)
        body = 'email,name\nstudent1@university.edu,Jane Doe\nstudent2@university.edu,John Roe\n'
        resp = HttpResponse(body, content_type='text/csv')
        resp['Content-Disposition'] = 'attachment; filename="bulk-invite-template.csv"'
        return resp


def _read_emails(upload):
    """Return [(row_number, email)] from a csv or xlsx upload. Row 1 is treated as a header
    when it contains an 'email' column; otherwise the first column is used."""
    name = (upload.name or '').lower()
    if name.endswith('.xlsx') or name.endswith('.xls'):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValueError('Spreadsheet support needs openpyxl installed; upload a .csv instead.')
        wb = load_workbook(upload, read_only=True, data_only=True)
        ws = wb.active
        grid = [[('' if c is None else str(c)) for c in row] for row in ws.iter_rows(values_only=True)]
    else:
        try:
            text = upload.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            raise ValueError('Could not read the file as UTF-8 text.')
        grid = [row for row in csv.reader(io.StringIO(text))]

    if not grid:
        return []

    header = [c.strip().lower() for c in grid[0]]
    if 'email' in header:
        col = header.index('email')
        data = grid[1:]
        start = 2
    else:
        col = 0
        data = grid
        start = 1

    out = []
    for offset, row in enumerate(data):
        value = row[col].strip() if col < len(row) else ''
        if value:
            out.append((start + offset, value))
    return out


# ---- registration link ----------------------------------------------------

class InstructorRegistrationLinkView(APIView):
    permission_classes = [IsAuthenticated, IsInstructor]

    def post(self, request, cohort_id):
        cohort = _cohort_for(request, cohort_id)
        regenerate = bool(request.data.get('regenerate'))
        if regenerate or not cohort.registration_token:
            cohort.registration_token = secrets.token_urlsafe(16)
            cohort.save(update_fields=['registration_token'])
        url = _registration_url(request, cohort.registration_token)
        return Response({'token': cohort.registration_token, 'url': url, 'enabled': True})

    def delete(self, request, cohort_id):
        cohort = _cohort_for(request, cohort_id)
        cohort.registration_token = ''
        cohort.save(update_fields=['registration_token'])
        return Response({'token': None, 'url': None, 'enabled': False})


# ---- extend round ---------------------------------------------------------

class InstructorExtendRoundView(APIView):
    permission_classes = [IsAuthenticated, IsInstructor]

    def post(self, request, cohort_id, n):
        cohort = _cohort_for(request, cohort_id)
        if not 1 <= n <= ADMIN_TOTAL_ROUNDS:
            return Response({'detail': 'No such round.'}, status=400)
        try:
            days = int(request.data.get('days'))
        except (TypeError, ValueError):
            return Response({'detail': 'days must be a whole number.'}, status=400)
        if days <= 0:
            return Response({'detail': 'days must be positive.'}, status=400)
        ext = dict(cohort.round_extensions or {})
        ext[str(n)] = int(ext.get(str(n), 0)) + days
        cohort.round_extensions = ext
        cohort.save(update_fields=['round_extensions'])
        current = max((r.current_week for r in Run.objects.filter(team__cohort=cohort)), default=1)
        return Response({'ok': True, 'rounds': _rounds(cohort, current)})


# ---- move enrollment ------------------------------------------------------

class InstructorMoveEnrollmentView(APIView):
    permission_classes = [IsAuthenticated, IsInstructor]

    def post(self, request, cohort_id, enrollment_id):
        cohort = _cohort_for(request, cohort_id)
        enrollment = get_object_or_404(Enrollment, id=enrollment_id, cohort=cohort)
        try:
            firm_number = int(request.data.get('firm_number'))
        except (TypeError, ValueError):
            return Response({'detail': 'firm_number is required.'}, status=400)

        if firm_number == 0:
            with transaction.atomic():
                _sync_team_membership(cohort, enrollment.student, None)
                enrollment.team = None
                enrollment.save(update_fields=['team'])
            return Response({'ok': True, 'firm': None})

        _, numbering = _firm_numbering(cohort)
        target = next((tid for tid, num in numbering.items() if num == firm_number), None)
        if target is None:
            return Response({'detail': 'No such firm.'}, status=400)
        team = Team.objects.get(id=target)
        with transaction.atomic():
            _sync_team_membership(cohort, enrollment.student, team)
            enrollment.team = team
            enrollment.save(update_fields=['team'])
        return Response({'ok': True, 'firm': team.name})


# ---- co-faculty -----------------------------------------------------------

class InstructorCoFacultyView(APIView):
    """Add a co-teacher to a cohort.

    Faculty could only be attached by an admin at provisioning, so bringing in a
    TA or a co-lecturer mid-term meant going back to whoever holds admin. An
    instructor already trusted with a cohort can staff it.

    If the address has no account yet, one is created and the person is emailed
    a link to choose their own password — the same flow the admin console uses.
    """

    permission_classes = [IsAuthenticated, IsInstructor]

    def post(self, request, cohort_id):
        cohort = _cohort_for(request, cohort_id)
        email = (request.data.get('email') or '').strip().lower()
        first_name = (request.data.get('first_name') or '').strip()
        last_name = (request.data.get('last_name') or '').strip()

        errors = {}
        if not email or '@' not in email:
            errors['email'] = 'Enter a valid email address.'
        if errors:
            return Response({'detail': 'Some fields need attention.', 'errors': errors}, status=400)

        existing = User.objects.filter(username__iexact=email).first()
        invited = False

        if existing:
            if existing.role != UserRole.INSTRUCTOR:
                return Response({
                    'detail': 'Some fields need attention.',
                    'errors': {'email': 'That email belongs to a student account.'},
                }, status=400)
            if cohort.instructors.filter(pk=existing.pk).exists():
                return Response({
                    'detail': 'Some fields need attention.',
                    'errors': {'email': 'They already teach this simulation.'},
                }, status=400)
            user = existing
        else:
            if not first_name:
                return Response({
                    'detail': 'Some fields need attention.',
                    'errors': {'first_name': 'New instructors need a first name.'},
                }, status=400)
            user = User.objects.create_user(
                username=email, email=email, password=None,
                first_name=first_name, last_name=last_name, role=UserRole.INSTRUCTOR,
            )
            invited = True

        cohort.instructors.add(user)

        sent, url, error = (True, '', '')
        if invited:
            sent, url, error = send_faculty_invite(user, invited_by=request.user)

        return Response({
            'id': user.id,
            'username': user.username,
            'name': (user.get_full_name() or '').strip() or user.username,
            'created': invited,
            'invite_sent': sent if invited else None,
            **({'set_password_url': url, 'send_error': error} if invited and not sent else {}),
        }, status=201)


class InstructorCoFacultyDetailView(APIView):
    """Remove a co-teacher from a cohort.

    The account is untouched — this only unlinks them from this cohort, and only
    while somebody is left teaching it. A cohort with no instructors is
    unreachable: every instructor endpoint scopes by who teaches it, so nobody
    could add one back.
    """

    permission_classes = [IsAuthenticated, IsInstructor]

    def delete(self, request, cohort_id, user_id):
        cohort = _cohort_for(request, cohort_id)
        user = get_object_or_404(User, id=user_id)
        if not cohort.instructors.filter(pk=user.pk).exists():
            return Response({'detail': 'They do not teach this simulation.'}, status=404)
        if cohort.instructors.count() <= 1:
            return Response({
                'detail': (
                    'This is the only instructor on the simulation. Add someone else '
                    'before removing them, or nobody will be able to reach it.'
                ),
            }, status=409)

        cohort.instructors.remove(user)
        return Response({'ok': True, 'removed': (user.get_full_name() or '').strip() or user.username})


# ---- firms: create, delete ------------------------------------------------

MAX_FIRMS = 20


class InstructorFirmsView(APIView):
    """Add a firm to a cohort.

    Firms could only be created at provisioning, by an admin. A term does not
    hold still — students arrive late, a firm of five is better split than
    stretched — so faculty create them here, with the run the engine needs
    attached at the same moment.
    """

    permission_classes = [IsAuthenticated, IsInstructor]

    def post(self, request, cohort_id):
        cohort = _cohort_for(request, cohort_id)
        existing = Team.objects.filter(cohort=cohort)
        if existing.count() >= MAX_FIRMS:
            return Response(
                {'detail': f'A cohort holds at most {MAX_FIRMS} firms.'}, status=400,
            )

        name = (request.data.get('name') or '').strip()
        if not name:
            # Next free "Team N" — counting is not enough, because deleting
            # Team 2 of three would otherwise propose a name already in use.
            taken = set(existing.values_list('name', flat=True))
            n = 1
            while f'Team {n}' in taken:
                n += 1
            name = f'Team {n}'
        elif len(name) > 255:
            return Response({'detail': 'Keep the firm name under 255 characters.'}, status=400)
        elif existing.filter(name__iexact=name).exists():
            return Response(
                {'detail': f'This cohort already has a firm called {name}.'}, status=400,
            )

        with transaction.atomic():
            team = Team.objects.create(cohort=cohort, name=name)
            Run.objects.create(team=team)

        _, numbering = _firm_numbering(cohort)
        return Response(
            {'id': team.id, 'name': team.name, 'number': numbering.get(team.id)}, status=201,
        )


class InstructorFirmDetailView(APIView):
    """Delete a firm.

    Deliberately obstructive. Run is a OneToOne on Team with CASCADE, so
    deleting a firm takes its run, its week instances, its submissions and its
    score records with it — a term's graded work, gone on one click and not
    recoverable from the UI. So a firm is only deletable while it is genuinely
    empty: nobody in it, and nothing ever submitted.
    """

    permission_classes = [IsAuthenticated, IsInstructor]

    def delete(self, request, cohort_id, firm_number):
        cohort = _cohort_for(request, cohort_id)
        _, numbering = _firm_numbering(cohort)
        team_id = next((tid for tid, num in numbering.items() if num == firm_number), None)
        if team_id is None:
            return Response({'detail': 'No such firm in this cohort.'}, status=404)
        team = Team.objects.get(id=team_id)

        members = team.members.count()
        if members:
            return Response({
                'detail': (
                    f'{team.name} still has {members} student'
                    f'{"" if members == 1 else "s"} in it. Move them out first — '
                    'deleting a firm would delete their work with it.'
                ),
            }, status=409)

        run = Run.objects.filter(team=team).first()
        if run:
            worked = run.week_instances.filter(
                status__in=(WeekInstanceStatus.SUBMITTED, WeekInstanceStatus.SCORED)
            ).count()
            if worked:
                return Response({
                    'detail': (
                        f'{team.name} has {worked} submitted round'
                        f'{"" if worked == 1 else "s"} on record. That work cannot be '
                        'recovered once the firm is deleted, so this is blocked.'
                    ),
                }, status=409)

        name = team.name
        team.delete()  # takes the empty run with it
        return Response({'ok': True, 'deleted': name})


# ---- setup test team ------------------------------------------------------

class InstructorSetupTestTeamView(APIView):
    """Provision test-N@mailinator.com accounts and round-robin enroll them across firms.
    No executive seats, since students only belong to a firm. All share TEST_PASSWORD."""

    permission_classes = [IsAuthenticated, IsInstructor]

    def post(self, request, cohort_id):
        cohort = _cohort_for(request, cohort_id)
        teams, _ = _firm_numbering(cohort)
        if not teams:
            return Response({'detail': 'Create firms before provisioning a test team.'}, status=409)
        try:
            per_firm = int(request.data.get('per_firm') or 4)
        except (TypeError, ValueError):
            per_firm = 4
        per_firm = max(1, min(per_firm, 20))

        total = per_firm * len(teams)
        created = reused = 0
        accounts = []
        with transaction.atomic():
            for i in range(total):
                email = f'test-{i + 1}@mailinator.com'
                student, was_created = User.objects.get_or_create(
                    username=email,
                    defaults={'email': email, 'role': UserRole.STUDENT},
                )
                if was_created:
                    student.set_password(TEST_PASSWORD)
                    student.save(update_fields=['password'])
                    created += 1
                else:
                    reused += 1
                team = teams[i % len(teams)]
                Enrollment.objects.update_or_create(
                    cohort=cohort, student=student, defaults={'team': team},
                )
                _sync_team_membership(cohort, student, team)
                accounts.append({'email': email, 'firm': team.name})

        return Response({
            'total': total,
            'created': created,
            'reused': reused,
            'firms': len(teams),
            'password': TEST_PASSWORD,
            'accounts': accounts,
        })

class InstructorEnrollmentPaidView(APIView):
    """Mark a single enrollment paid or unpaid. Body: {"paid": true|false}."""

    permission_classes = [IsAuthenticated, IsInstructor]

    def post(self, request, cohort_id, enrollment_id):
        cohort = _cohort_for(request, cohort_id)
        enrollment = get_object_or_404(Enrollment, id=enrollment_id, cohort=cohort)
        enrollment.paid = bool(request.data.get('paid', True))
        enrollment.save(update_fields=['paid'])
        return Response({'enrollment_id': enrollment.id, 'paid': enrollment.paid})


class InstructorBulkPaidView(APIView):
    """Mark many enrollments paid or unpaid at once.

    Body: {"paid": true|false, "enrollment_ids": [..]} — omit enrollment_ids to
    apply to every enrollment in the cohort. Ids outside the cohort and
    non-numeric entries are ignored.
    """

    permission_classes = [IsAuthenticated, IsInstructor]

    def post(self, request, cohort_id):
        cohort = _cohort_for(request, cohort_id)
        paid = bool(request.data.get('paid', True))
        qs = Enrollment.objects.filter(cohort=cohort)
        ids = request.data.get('enrollment_ids')
        if ids is not None:
            clean = []
            for value in (ids if isinstance(ids, (list, tuple)) else [ids]):
                try:
                    clean.append(int(value))
                except (TypeError, ValueError):
                    continue
            qs = qs.filter(id__in=clean)
        updated = qs.update(paid=paid)
        return Response({'updated': updated, 'paid': paid})


class InstructorInsightsView(APIView):
    """Everything the Insights page renders, in one call.

    Per firm: graded totals per week, dimension sums, momentum (last graded
    round vs the one before), trap-flag count, and advisor usage. Cohort-wide:
    a firm x round status matrix (scored total / submitted / open), dimension
    averages across all graded weeks, and benchmark payloads when they exist.
    Instructor-side, so trap flags are fair game here — students never see them.
    """

    permission_classes = [IsAuthenticated, IsInstructor]

    def get(self, request, cohort_id):
        from scoring.models import Benchmark
        from scoring.services import student_benchmark_payload
        from weeks.models import WeekInstance, WeekInstanceStatus

        cohort = _cohort_for(request, cohort_id)
        teams = list(Team.objects.filter(cohort=cohort).order_by('id'))
        runs = {r.team_id: r for r in Run.objects.filter(team__cohort=cohort)}
        current_week = max((r.current_week for r in runs.values()), default=1) or 1

        instances = (
            WeekInstance.objects
            .filter(run__team__cohort=cohort)
            .select_related('score', 'run')
            .order_by('week_number')
        )
        by_team = {}
        for inst in instances:
            by_team.setdefault(inst.run.team_id, []).append(inst)

        advisor_usage = {
            row['team_id']: row
            for row in AdvisorSession.objects
            .for_cohort(cohort)
            # A session hangs off one path or the other, so the team lives on
            # whichever of the two is set.
            .annotate(team_id=Coalesce('conversation__run__team_id', 'group_session__run__team_id'))
            .values('team_id')
            .annotate(hours=Count('id'), billed=Sum(BILLED))
        }

        dims = ['strategic_judgment', 'execution_consequence', 'coherence', 'deliverable_quality']
        dim_totals = {d: 0 for d in dims}
        # The mean alone hides the round that did the most work: a cohort where
        # half the firms scored +2 and half -2 reads as a flat 0. Keep every
        # value so the spread can be reported alongside it.
        dim_values = {d: [] for d in dims}
        dim_count = 0

        firms = []
        for index, team in enumerate(teams, start=1):
            weeks = []
            totals = []
            traps = 0
            for inst in by_team.get(team.id, []):
                score = getattr(inst, 'score', None)
                scored = inst.status == WeekInstanceStatus.SCORED and score is not None
                cell = {'week': inst.week_number, 'status': inst.status}
                if scored:
                    d = score.dimension_scores()
                    cell['total'] = sum(d.values())
                    cell['scores'] = d
                    totals.append(cell['total'])
                    for k, v in d.items():
                        dim_totals[k] += v
                        dim_values[k].append(v)
                    dim_count += 1
                if score is not None:
                    traps += len(score.auto_components.get('trap_flags', []))
                weeks.append(cell)
            usage = advisor_usage.get(team.id, {})
            momentum = (totals[-1] - totals[-2]) if len(totals) >= 2 else None
            firms.append({
                'number': index,
                'name': team.name,
                'members': team.members.count(),
                'weeks': weeks,
                'graded_count': len(totals),
                'total_score': sum(totals),
                'average': round(sum(totals) / len(totals), 1) if totals else 0,
                'last_total': totals[-1] if totals else None,
                'momentum': momentum,
                'trap_flags': traps,
                'advisor_hours': usage.get('hours', 0) or 0,
                'advisor_billed': usage.get('billed', 0) or 0,
            })
        firms.sort(key=lambda f: (-f['total_score'], f['number']))

        # Same withholding rule as the Benchmarks screen: a table missing a
        # firm ranks the rest wrongly, so it is not shown until the round is
        # fully graded. benchmark_status below says who is outstanding.
        from scoring.services import benchmark_pending as _pending
        benchmarks = [
            student_benchmark_payload(b)
            for b in Benchmark.objects.filter(cohort=cohort).order_by('after_week')
            if not _pending(cohort, b.after_week)
        ]

        # A benchmark round only publishes once every playing firm is graded.
        # Without this the gate is indistinguishable from a bug: the instructor
        # grades a Week 4 and no standings appear, with nothing on screen
        # saying who is still outstanding.
        from scoring.config import BENCHMARK_PHASE_WEEKS
        from scoring.services import benchmark_pending
        benchmark_status = None
        due = [w for w in BENCHMARK_PHASE_WEEKS if w <= current_week]
        if due:
            after_week = max(due)
            pending = benchmark_pending(cohort, after_week)
            benchmark_status = {
                'after_week': after_week,
                'published': any(b['after_week'] == after_week for b in benchmarks),
                'pending_firms': [t.name for t in pending],
            }

        return Response({
            'current_week': current_week,
            'total_rounds': ADMIN_TOTAL_ROUNDS,
            'firms': firms,
            'benchmark_status': benchmark_status,
            'dimension_averages': {d: round(dim_totals[d] / dim_count, 1) if dim_count else 0 for d in dims},
            # Mean, range and how many firm-weeks landed either side of zero.
            # `split` is the tell: a dimension with firms on both sides is one
            # where the round separated them, whatever the mean says.
            'dimension_spread': {
                d: {
                    'min': min(dim_values[d]) if dim_values[d] else 0,
                    'max': max(dim_values[d]) if dim_values[d] else 0,
                    'above': sum(1 for v in dim_values[d] if v > 0),
                    'below': sum(1 for v in dim_values[d] if v < 0),
                    'count': len(dim_values[d]),
                }
                for d in dims
            },
            'graded_weeks_total': dim_count,
            'benchmarks': benchmarks,
        })


class InstructorFirmInsightsView(APIView):
    """Deep dive for one firm: members with advisor usage, week-by-week graded
    record with trap flags, and a curated snapshot of the run's engine state —
    stakeholder relationships, through-lines, gates, accumulated scores, the
    coherence anchor, and recent decision history. Backs the firm dashboards
    page. Nested state blobs are passed through as-is; the client renders
    them defensively."""

    permission_classes = [IsAuthenticated, IsInstructor]

    def get(self, request, cohort_id, firm_number):
        from weeks.models import WeekInstance, WeekInstanceStatus

        cohort = _cohort_for(request, cohort_id)
        teams, numbering = _firm_numbering(cohort)
        team = next((t for t in teams if numbering.get(t.id) == firm_number), None)
        if team is None:
            return Response({'detail': 'No such firm in this cohort.'}, status=404)
        run = Run.objects.filter(team=team).first()

        member_usage = {
            row['student_id']: row
            for row in AdvisorSession.objects.for_team(team)
            .values('student_id')
            .annotate(
                hours=Count('id'), billed=Sum(BILLED),
                group_hours=GROUP_HOURS, group_billed=GROUP_BILLED,
            )
        }
        members = [
            {
                'id': u.id,
                'name': (u.get_full_name() or u.username),
                'email': u.email,
                'advisor_hours': member_usage.get(u.id, {}).get('hours', 0) or 0,
                'advisor_billed': member_usage.get(u.id, {}).get('billed', 0) or 0,
                'group_hours': member_usage.get(u.id, {}).get('group_hours', 0) or 0,
                'group_billed': member_usage.get(u.id, {}).get('group_billed', 0) or 0,
            }
            for u in team.members.all()
        ]

        weeks = []
        if run:
            for inst in WeekInstance.objects.filter(run=run).select_related('score').order_by('week_number'):
                score = getattr(inst, 'score', None)
                row = {'week': inst.week_number, 'status': inst.status}
                if score is not None:
                    row['trap_flags'] = score.auto_components.get('trap_flags', [])
                    if inst.status == WeekInstanceStatus.SCORED:
                        dims = score.dimension_scores()
                        row['scores'] = dims
                        row['total'] = sum(dims.values())
                weeks.append(row)

        state = (run.state or {}) if run else {}
        history = state.get('decision_history') or []
        return Response({
            'number': firm_number,
            'name': team.name,
            'current_week': run.current_week if run else None,
            'members': members,
            'weeks': weeks,
            'state': {
                'coherence_anchor': state.get('coherence_anchor', ''),
                'accumulated_scores': state.get('accumulated_scores', {}),
                'relationships': state.get('relationships', {}),
                'through_lines': state.get('through_lines', {}),
                'gates': state.get('gates', {}),
                'flags': state.get('flags', {}),
                'recent_decisions': history[-6:],
                'decision_count': len(history),
            },
        })


class InstructorMimicRunView(APIView):
    """The student game screen for one firm, as the instructor's read-only
    mimic. Returns the same payload shape as the student /run/ endpoint plus
    the firm's submission when one exists — but built without view_briefing,
    so peeking never mutates the week instance's status the way a real
    student opening the briefing does."""

    permission_classes = [IsAuthenticated, IsInstructor]

    def get(self, request, cohort_id, firm_number):
        from advisors.models import AdvisorDefinition
        from core.models import RunStatus
        from weeks.models import WeekInstanceStatus
        from weeks.registry import registry

        from . import serialize
        from .views import briefing_for

        cohort = _cohort_for(request, cohort_id)
        teams, numbering = _firm_numbering(cohort)
        team = next((t for t in teams if numbering.get(t.id) == firm_number), None)
        if team is None:
            return Response({'detail': 'No such firm in this cohort.'}, status=404)
        run = Run.objects.filter(team=team).first()
        if run is None:
            return Response({'detail': 'No run yet for this firm.'}, status=404)

        module = registry.get(run.current_week)
        tier = cohort.tier
        instance = run.week_instances.filter(week_number=run.current_week).first()
        if instance is not None:
            week = serialize.week_status(instance)
        else:
            # The firm hasn't opened this week yet — synthesize the untouched state.
            week = {'week_number': run.current_week, 'status': 'BRIEFING', 'submitted': False, 'scored': False}

        submission = None
        sub = getattr(instance, 'submitted_payload', None) if instance else None
        if sub is not None:
            submission = {
                'structured_payload': sub.structured_payload,
                'deliverable_text': sub.deliverable_text,
            }

        week14 = run.week_instances.filter(week_number=14).first()
        debrief_ready = bool(
            run.status == RunStatus.COMPLETE
            or (week14 and week14.status in (WeekInstanceStatus.SUBMITTED, WeekInstanceStatus.SCORED))
        )
        debrief = None
        if debrief_ready:
            from engine.climax import generate_debrief
            try:
                debrief = generate_debrief(run.state)
            except Exception:
                debrief = None

        # The firm's graded record, shaped exactly like /student/performance/
        # so the mimic Performance section shows what students see.
        perf_rows = []
        for wi in run.week_instances.filter(status=WeekInstanceStatus.SCORED).select_related('score').order_by('week_number'):
            score = getattr(wi, 'score', None)
            if score is None:
                continue
            dims = score.dimension_scores()
            perf_rows.append({
                'week_number': wi.week_number,
                'scores': dims,
                'total': sum(dims.values()),
                'graded_at': score.graded_at.isoformat() if score.graded_at else None,
            })
        performance = {
            'weeks': perf_rows,
            'graded_count': len(perf_rows),
            'average': round(sum(r['total'] for r in perf_rows) / len(perf_rows), 1) if perf_rows else 0,
            'best': max((r['total'] for r in perf_rows), default=0),
        }

        return Response({
            'mimic': True,
            'firm': {'number': firm_number, 'name': team.name},
            'run': {
                'id': run.id,
                'current_week': run.current_week,
                'status': run.status,
                'tier_outcome': run.tier_outcome,
            },
            'week': week,
            'briefing': serialize.briefing_json(
                briefing_for(module, tier, run), instance.preamble if instance else ''
            ),
            'artifacts': serialize.artifacts_json(module.artifacts(tier)),
            'decision_spec': serialize.decision_spec_json(module.decision_spec(tier)),
            'advisors': [serialize.advisor_json(a) for a in AdvisorDefinition.objects.filter(active=True)],
            'debrief_available': debrief_ready,
            'debrief': debrief,
            'performance': performance,
            'submission': submission,
        })


class InstructorStudentDetailView(APIView):
    """Everything about one enrolled student: profile, enrollment and access
    state, and their advisor session history with charges."""

    permission_classes = [IsAuthenticated, IsInstructor]

    def get(self, request, cohort_id, enrollment_id):
        cohort = _cohort_for(request, cohort_id)
        enrollment = get_object_or_404(
            Enrollment.objects.select_related('student', 'team'), id=enrollment_id, cohort=cohort
        )
        user = enrollment.student
        teams, numbering = _firm_numbering(cohort)
        firm_number = numbering.get(enrollment.team_id) if enrollment.team_id else None

        sessions = (
            AdvisorSession.objects
            .filter(enrollment=enrollment)
            .select_related('conversation__advisor', 'group_session')
            .order_by('-started_at')
        )
        session_rows = [
            {
                'mode': 'group' if s.group_session_id else 'solo',
                'advisor': (
                    s.conversation.advisor.name if s.conversation_id
                    else f'War room · {s.advisor_count} advisors'
                ),
                'week': (
                    s.conversation.week_number if s.conversation_id
                    else (s.group_session.week_number if s.group_session_id else None)
                ),
                'started_at': s.started_at.isoformat(),
                'rate': s.hourly_rate,
                'advisor_count': s.advisor_count,
                'billed': s.billed,
            }
            for s in sessions
        ]
        return Response({
            'enrollment_id': enrollment.id,
            'user': {
                'id': user.id,
                'name': (user.get_full_name() or user.username),
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email or user.username,
                'username': user.username,
                'date_joined': user.date_joined.isoformat() if user.date_joined else None,
                'last_login': user.last_login.isoformat() if user.last_login else None,
            },
            'enrollment': {
                'firm': enrollment.team.name if enrollment.team else None,
                'firm_number': firm_number,
                'paid': enrollment.paid,
                'paid_at': enrollment.paid_at.isoformat() if enrollment.paid_at else None,
                'blocked': enrollment.blocked,
                'blocked_at': enrollment.blocked_at.isoformat() if enrollment.blocked_at else None,
                'amount_due': enrollment.amount_due,
            },
            'advisor': {
                'hours': len(session_rows),
                # Per-row billed, not the raw rate: a war-room hour is charged
                # once per advisor seated in it.
                'billed': sum(r['billed'] for r in session_rows),
                'group_hours': sum(1 for r in session_rows if r['mode'] == 'group'),
                'group_billed': sum(r['billed'] for r in session_rows if r['mode'] == 'group'),
                'sessions': session_rows,
            },
        })


class InstructorStudentBlockView(APIView):
    """Block or unblock a student's access to this cohort.

    Body: {"blocked": true|false}. Blocking flips the enrollment flag the
    student portal already respects (their cohort card gates gameplay and
    shows the paused notice); it does not delete anything and is instantly
    reversible.
    """

    permission_classes = [IsAuthenticated, IsInstructor]

    def post(self, request, cohort_id, enrollment_id):
        cohort = _cohort_for(request, cohort_id)
        enrollment = get_object_or_404(Enrollment, id=enrollment_id, cohort=cohort)
        blocked = bool(request.data.get('blocked', True))
        enrollment.blocked = blocked
        enrollment.blocked_at = dj_timezone.now() if blocked else None
        enrollment.save(update_fields=['blocked', 'blocked_at'])
        return Response({'enrollment_id': enrollment.id, 'blocked': enrollment.blocked})


class InstructorStudentResetPasswordView(APIView):
    """Set a fresh temporary password for a student and return it ONCE.

    Classroom-tool pattern: the instructor reads the temp password to the
    student, who logs in with it. The password is generated server-side
    (16 chars, unambiguous alphabet), never stored in plaintext, and never
    retrievable again — refreshing this page will not show it twice.
    Note: JWTs already issued remain valid until they expire; blocking is
    the tool for cutting off live access.
    """

    permission_classes = [IsAuthenticated, IsInstructor]

    ALPHABET = 'abcdefghjkmnpqrstuvwxyzACDEFGHJKLMNPQRSTUVWXYZ2345679'

    def post(self, request, cohort_id, enrollment_id):
        cohort = _cohort_for(request, cohort_id)
        enrollment = get_object_or_404(Enrollment.objects.select_related('student'), id=enrollment_id, cohort=cohort)
        user = enrollment.student
        temp_password = ''.join(secrets.choice(self.ALPHABET) for _ in range(16))
        user.set_password(temp_password)
        user.save(update_fields=['password'])
        return Response({'enrollment_id': enrollment.id, 'temp_password': temp_password})

class InstructorPeopleView(APIView):
    """Students and co-faculty across every cohort this instructor teaches.

    Backs the instructor workspace's Students and Faculty tabs — one roster that
    spans all of the instructor's simulations, rather than the per-cohort detail
    view. Students carry which sim and firm they're in plus advisor usage; each
    appears once per enrollment (a student in two of this instructor's cohorts
    shows up for each)."""

    permission_classes = [IsAuthenticated, IsInstructor]

    def get(self, request):
        cohorts = list(request.user.instructed_cohorts.all().order_by('name'))

        students = []
        for cohort in cohorts:
            _, numbering = _firm_numbering(cohort)
            enrollments = Enrollment.objects.filter(cohort=cohort).select_related('student', 'team')
            usage = _advisor_usage_by_student(cohort)
            for row in _student_rows(enrollments, numbering, usage):
                students.append({**row, 'cohort': cohort.name, 'cohort_id': cohort.id})

        # Co-faculty: every instructor on these cohorts, with the shared ones and
        # a flag for the requester so the UI can mark "you".
        faculty_map = {}
        for cohort in cohorts:
            for instr in cohort.instructors.all():
                entry = faculty_map.setdefault(instr.id, {
                    'id': instr.id,
                    'name': instr.get_full_name() or instr.username,
                    'email': instr.email or '',
                    'is_you': instr.id == request.user.id,
                    'cohorts': [],
                })
                entry['cohorts'].append(cohort.name)
        faculty = sorted(faculty_map.values(), key=lambda f: (not f['is_you'], f['name'].lower()))

        return Response({'students': students, 'faculty': faculty})
